"""
bank_reconciliation.py
======================
Reusable bank-statement <-> aging reconciliation.

Give it:
  * a bank statement workbook (one or more sheets of credit transactions), and
  * an aging / outstanding-receivables workbook,
and it writes one reconciliation workbook per bank sheet, each containing:

    Matched            - confident policy-level matches (with a Score column)
    Additional Matches - lower-confidence name/NIC/plate leads (verify first)
    Payer Clusters     - unmatched money grouped by payer, for reconciliation

It reproduces the schema of the reference workbooks (Bank_Matching.xlsx,
Recon_Bank_of_Ceylon.xlsx, Recon_DFCC_Bank.xlsx).

--------------------------------------------------------------------------
HOW TO REUSE ON NEW FILES
--------------------------------------------------------------------------
1. Point BANK_FILE / AGING_FILE / OUTPUT_DIR below at your files.
2. If a new statement has different column positions, add/adjust a mapping in
   BANK_SHEET_LAYOUTS (date / description / amount / optional reference).
3. If the aging file uses different header names, adjust AGING_COLS.
4. Run:   python bank_reconciliation.py

Requires:  openpyxl, rapidfuzz      (pip install openpyxl rapidfuzz)
--------------------------------------------------------------------------
"""

import re
from collections import defaultdict, Counter
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz
import openpyxl
import streamlit as st

from io import BytesIO

# ==========================================================================
# CONFIG  — edit these for your environment / new files
# ==========================================================================
# BANK_FILE  = "/mnt/user-data/uploads/Bank_Statement.xlsx"
# AGING_FILE = "/mnt/user-data/uploads/Aging_with_NIC.xlsx"
# AGING_SHEET = "SQL Results"
# OUTPUT_DIR = "/mnt/user-data/outputs"




st.title("Bank Reconciliation")

# Upload files
BANK_FILE = st.file_uploader(
    "Upload Bank Statement",
    type=["xlsx"]
)

AGING_FILE = st.file_uploader(
    "Upload Aging File",
    type=["xlsx"]
)

# Sheet name
AGING_SHEET = st.text_input(
    "Aging Sheet Name",
    value="SQL Results"
)

if BANK_FILE is None:
    st.info("Please upload the Bank Statement.")
    st.stop()

if AGING_FILE is None:
    st.info("Please upload the Aging file.")
    st.stop()

# Safe to continue from here
# ag = Aging(AGING_FILE, AGING_SHEET)

# Aging header names -> internal keys.  Change the *values* if your aging file
# uses different column titles.
AGING_COLS = {
    "policy":   "POLICY NO",
    "doc":      "DOC_NO",
    "org":      "ORG_LC_AMT",
    "recd":     "AMT Recd",
    "bal":      "BAL_LC_AMT",
    "ageing":   "AGEING_DUE_DAYS",
    "insured":  "INSURED NAME",
    "customer": "CUST_NAME",
    "producer": "PRODUCER_NAME",
    "division": "Divn Name",
    "nic":      "POLH_CIVIL_ID",
    "vehicle":  "VEH_REG_NO",
}

# Per-sheet column layout of the bank statement (0-based indexes, and the first
# data row).  Add an entry for any new sheet.  'ref' = -1 means "no ref column".
BANK_SHEET_LAYOUTS = {
    "Commertial Bank": {"date": 0, "desc": 1, "amount": 2, "ref": -1, "start": 2,
                        "out": "Recon_Commercial_Bank.xlsx", "label": "Commercial Bank"},
    "Bank of Ceylon":  {"date": 0, "desc": 2, "amount": 3, "ref": -1, "start": 3,
                        "out": "Recon_Bank_of_Ceylon.xlsx", "label": "Bank of Ceylon"},
    "DFCC Bank":       {"date": 0, "desc": 3, "amount": 4, "ref": 2, "start": 2,
                        "out": "Recon_DFCC_Bank.xlsx", "label": "DFCC Bank"},
}

# ==========================================================================
# SCORING WEIGHTS  (change these to re-tune the matcher)
# ==========================================================================
W_POLICY  = 100   # exact policy or document number
W_NIC9    = 60    # 9-digit NIC (+ V/X)
W_PLATE   = 45    # vehicle plate core
W_NIC12   = 32    # 12-digit NIC
W_NAME    = 15    # per shared insured name token ...
W_NAME_CAP = 30   # ... capped at this
W_AMOUNT  = 25    # amount corroborates Original / Outstanding / Net-due
AMT_TOL   = 0.02  # 2% amount tolerance
MATCH_MIN = 45    # minimum score to appear in the Matched tab
NAME_FUZZ = 90    # rapidfuzz ratio to treat two name tokens as the same

# ==========================================================================
# Helpers
# ==========================================================================
def fnum(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None

def norm(s):
    return re.sub(r"[^A-Za-z0-9]", "", str(s or "")).upper()

STOP = set(
    "THE AND FOR LTD PVT PLC LIMITED PRIVATE COMPANY MRS MISS MST INSURANCE INSUR "
    "INSURENCE INSUARANCE INSUARENCE PREMIUM POLICY CAR VEHICLE MOTOR CASH CHEQUE CHQ "
    "DEPOSIT PAYMENT TRANSFER CREDIT INWARD OUTWARD FUND BANK BRANCH COLLECTION DAILY "
    "SALES ORIENT TRAVEL LIFE GENERAL FIRST".split()
)
COMMON = set(
    "SILVA PERERA FERNANDO KUMARA BANDARA MOHAMED DIAS JAYASINGHE GUNAWARDANA "
    "RATHNAYAKE DISSANAYAKE SENANAYAKE WICKRAMASINGHE JAYAWARDENA WIJESINGHE RAJAPAKSE "
    "KARUNARATNE SANJEEWA CHANDANA NISHANTHA PRIYANKARA WIJERATNE".split()
)
def mtok(s):
    return [t for t in re.findall(r"[A-Za-z]{3,}", str(s or "").upper()) if t not in STOP]

POL_RE = re.compile(r"P\s*/?\s*\d+\s*/\s*\d+\s*/\s*\d+\s*/\s*\d+|\bP\d{12,}", re.I)
DUW_RE = re.compile(r"DUW\s*[-/]?\s*\d{2,}", re.I)
NIC_RE = re.compile(r"\b(\d{9}[VXvx]|\d{12})\b")

def bank_plates(desc):
    U = re.sub(r"TR\d{5,}", "", norm(desc))
    out = set()
    for m in re.finditer(r"([A-Z]{2,3})(\d{3,4})", U):
        pc = m.group(1) + m.group(2)
        if not pc.startswith("TR"):
            out.add(pc)
    return out


# ==========================================================================
# Aging index
# ==========================================================================
class Aging:
    def __init__(self, path, sheet):
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheet]
        data = list(ws.iter_rows(values_only=True))
        hdr = list(data[0])
        self.rows = [r for r in data[1:] if any(c is not None for c in r)]
        H = {h: i for i, h in enumerate(hdr)}
        self.c = {k: H[v] for k, v in AGING_COLS.items()}

        self.pol_idx = defaultdict(list)
        self.doc_idx = defaultdict(list)
        self.nic9_idx = defaultdict(list)
        self.nic12_idx = defaultdict(list)
        self.plate_idx = defaultdict(list)
        self.name_idx = defaultdict(set)     # token -> row indexes (for candidate gathering)
        self.itok = []                        # per-row insured token set
        self.org_ct = Counter()
        self.bal_ct = Counter()
        c = self.c
        for i, r in enumerate(self.rows):
            if norm(r[c["policy"]]):
                self.pol_idx[norm(r[c["policy"]])].append(i)
            if norm(r[c["doc"]]):
                self.doc_idx[norm(r[c["doc"]])].append(i)
            nic = norm(r[c["nic"]])
            if re.fullmatch(r"\d{9}[VX]", nic):
                self.nic9_idx[nic].append(i)
            elif re.fullmatch(r"\d{12}", nic):
                self.nic12_idx[nic].append(i)
            for m in re.finditer(r"([A-Z]{2,3})(\d{3,4})", norm(r[c["vehicle"]])):
                self.plate_idx[m.group(1) + m.group(2)].append(i)
            it = set(mtok(r[c["insured"]]))
            self.itok.append(it)
            for t in it:
                if len(t) >= 3:
                    self.name_idx[t].add(i)
            o, b = fnum(r[c["org"]]), fnum(r[c["bal"]])
            if o:
                self.org_ct[round(o, 2)] += 1
            if b and b > 0:
                self.bal_ct[round(b, 2)] += 1

    def veh_plates(self, i):
        return {m.group(1) + m.group(2)
                for m in re.finditer(r"([A-Z]{2,3})(\d{3,4})", norm(self.rows[i][self.c["vehicle"]]))}

    def amt_agree(self, paid, i):
        r = self.rows[i]; c = self.c
        o, b = fnum(r[c["org"]]), fnum(r[c["bal"]])
        nb = (o - fnum(r[c["recd"]])) if (o is not None and fnum(r[c["recd"]]) is not None) else None
        for label, v, ct in (("Original", o, self.org_ct),
                             ("Outstanding", b, self.bal_ct),
                             ("Net due", nb, None)):
            if v and paid and abs(paid - v) / max(abs(v), 1) <= AMT_TOL:
                freq = ct.get(round(v, 2), 1) if ct is not None else 1
                return (label, freq)
        return None


# ==========================================================================
# CONFIDENT scored matcher  (produces the "Matched" tab)
# ==========================================================================
SIGNAL_ORDER = ["Policy No", "Doc No", "NIC", "Vehicle", "Name", "NIC(12)", "Amount"]

def _name_score(bank_tokens, ins_tokens):
    """Shared tokens (exact or fuzzy) between narrative and insured name."""
    hits = 0
    for bt in bank_tokens:
        if bt in ins_tokens:
            hits += 1
        elif any(fuzz.ratio(bt, it) >= NAME_FUZZ for it in ins_tokens if abs(len(bt) - len(it)) <= 2):
            hits += 1
    return hits

def confident_match(ag, desc, paid):
    """Return (row_index, score, matched_on, amt_info) or None."""
    pols  = {norm(m.group()) for m in POL_RE.finditer(desc)}
    docs  = {norm(m.group()) for m in DUW_RE.finditer(desc)}
    nics  = [m.group(1).upper() for m in NIC_RE.finditer(desc.upper())]
    nic9  = {x for x in nics if re.fullmatch(r"\d{9}[VX]", x)}
    nic12 = {x for x in nics if re.fullmatch(r"\d{12}", x)}
    plts  = bank_plates(desc)
    d = re.sub(r"Bene/?OrgBK.*?(/|$)", " ", desc, flags=re.I)   # drop Orient's own account line
    ntok = set(mtok(d))

    cand = set()
    for p in pols:  cand |= set(ag.pol_idx.get(p, []))
    for x in docs:  cand |= set(ag.doc_idx.get(x, []))
    for x in nic9:  cand |= set(ag.nic9_idx.get(x, []))
    for x in nic12: cand |= set(ag.nic12_idx.get(x, []))
    for p in plts:  cand |= set(ag.plate_idx.get(p, []))
    for t in ntok:  cand |= ag.name_idx.get(t, set())

    best = None
    for i in cand:
        r = ag.rows[i]; c = ag.c
        s = 0; sig = []
        if norm(r[c["policy"]]) in pols:
            s += W_POLICY; sig.append("Policy No")
        elif norm(r[c["doc"]]) in docs:
            s += W_POLICY; sig.append("Doc No")
        if norm(r[c["nic"]]) in nic9:
            s += W_NIC9; sig.append("NIC")
        if plts and (plts & ag.veh_plates(i)):
            s += W_PLATE; sig.append("Vehicle")
        nh = _name_score(ntok, ag.itok[i])
        if nh:
            s += min(W_NAME * nh, W_NAME_CAP); sig.append("Name")
        if norm(r[c["nic"]]) in nic12:
            s += W_NIC12; sig.append("NIC(12)")
        am = ag.amt_agree(paid, i)
        if am:
            s += W_AMOUNT; sig.append("Amount")
        org = fnum(r[c["org"]])
        close = abs(paid - org) / max(org, 1) if (paid and org) else 9
        key = (s, -close, fnum(r[c["bal"]]) or 0)
        if best is None or key > best[0]:
            best = (key, i, s, sig, am)
    if best and best[2] >= MATCH_MIN:
        _, i, s, sig, am = best
        matched_on = " + ".join(x for x in SIGNAL_ORDER if x in sig)
        return (i, s, matched_on, am)
    return None


def paid_vs_invoice(paid, org):
    if not (paid and org):
        return ""
    if abs(paid - org) / max(org, 1) <= AMT_TOL:
        return "Full"
    if paid < org:
        return "Partial"
    if paid > 1.5 * org and (paid - org) > 1000:
        return "Over (bulk?)"
    return "Differs"


# ==========================================================================
# ADDITIONAL heuristic pass  (produces the "Additional Matches" tab)
#   distinctive insured surname/company + a rare matching amount
# ==========================================================================
REFUND = re.compile(r"refund|revers|mistaken|charge ?back", re.I)
TT_RE  = re.compile(r"\bTT\s*\d{3,}\b|\bCIB\b|INSURANCE BROK|B/?O\b.{0,25}BROK", re.I)
HASPOL = re.compile(r"P\s*/?\s*\d+\s*/\s*\d+\s*/\s*\d+\s*/\s*\d+|\bP\d{12,}", re.I)

def additional_matches(ag, unmatched):
    out = []
    for u in unmatched:
        desc, paid = u["desc"], u["amount"]
        if REFUND.search(desc) or TT_RE.search(desc) or HASPOL.search(desc) or not paid:
            continue
        d = re.sub(r"Bene/?OrgBK.*?(/|$)", " ", desc, flags=re.I)
        bt = set(mtok(d))
        cand = set()
        for t in bt:
            if len(t) >= 5 and t not in COMMON:
                cand |= ag.name_idx.get(t, set())
        best = None
        for i in cand:
            am = ag.amt_agree(paid, i)
            if not am or am[1] > 3:                      # amount must be rare (<=3 policies)
                continue
            shared = [t for t in (bt & ag.itok[i]) if len(t) >= 5 and t not in COMMON]
            if not shared:
                continue
            key = (len(shared), 3 - am[1], sum(len(t) for t in shared))
            if best is None or key > best[0]:
                best = (key, i, shared, am)
        if best:
            _, i, shared, am = best
            agrees = f"{am[0]} ({'unique' if am[1] == 1 else str(am[1]) + ' policies'})"
            out.append((u, i, "Name + Amount", ", ".join(sorted(shared)), agrees))
    return out

def nic_plate_singles(ag, unmatched, already):
    """Single-signal NIC / plate leads (unique policy only)."""
    out = []
    seen = set(already)
    for u in unmatched:
        key = (str(u["date"]), u["desc"], u["amount"])
        if key in seen:
            continue
        for m in NIC_RE.finditer(u["desc"].upper()):
            k = m.group(1)
            if k in ag.nic9_idx and len(ag.nic9_idx[k]) == 1:
                out.append((u, ag.nic9_idx[k][0], "NIC (single signal)", k, "-")); seen.add(key); break
            if k in ag.nic12_idx and len(ag.nic12_idx[k]) == 1:
                out.append((u, ag.nic12_idx[k][0], "NIC (single signal)", k, "-")); seen.add(key); break
    for u in unmatched:
        key = (str(u["date"]), u["desc"], u["amount"])
        if key in seen:
            continue
        U2 = re.sub(r"TR\d{5,}", "", norm(u["desc"]))
        for m in re.finditer(r"([A-Z]{2,3})(\d{3,4})", U2):
            pc = m.group(1) + m.group(2)
            if pc.startswith("TR"):
                continue
            if pc in ag.plate_idx and len(ag.plate_idx[pc]) == 1:
                out.append((u, ag.plate_idx[pc][0], "Vehicle plate (single signal)", pc, "-"))
                seen.add(key); break
    return out


# ==========================================================================
# PAYER CLUSTERS  (produces the "Payer Clusters" tab)
# ==========================================================================
BOIL = set(
    "THE AND FOR LTD PVT PLC LIMITED PRIVATE COMPANY MRS MISS BENE ORGBK ORIENT "
    "INSURANCE INSUR INSURENCE INSUARANCE INSUARENCE RTGS TRANSFER CHQ CHEQUE CASH "
    "DEPOSIT PAYMENT PYM EFT CEFT CEFTS OTC XXX REF INWARD OUTWARD FUND NEFT SWIFT TRF "
    "DEP INV SLIP VEHICLE RENEWAL PREM PREMIUM POL POLICY CAR MOTOR CREDIT DEBIT "
    "MOBILEBANKING BANKING COLLECTION DAILY SALES BRANCH NUMBER SENDER BIC BENEFICIARY "
    "DETAILS TRANSACTION ACCOUNT CRM CRDLESDEP REALTIME FIRST CMLOAN HNBF INSURANC SLIPS "
    "AECB AECG AECK AECN AECP AECT AECU LOLC IFT CTY CITILKLX CITIL BNK RMW OBFUND".split()
)
def _payer_tokens(desc):
    d = re.sub(r"Bene/?OrgBK.*?(/|$)", " ", desc, flags=re.I)
    d = re.sub(r"\b[A-Z]{4}LK[A-Z0-9]{2,6}\b", " ", d)
    d = re.sub(r"\b(FT|TT|CTY|IFT)\w{6,}", " ", d, flags=re.I)
    d = re.sub(r"\b[A-Z]*\d+[A-Z]*\b", " ", d, flags=re.I)
    toks = [t for t in re.findall(r"[A-Za-z]{3,}", d.upper()) if t not in BOIL]
    out = []
    for t in toks:
        if not out or out[-1] != t:
            out.append(t)
    return out

def _bankref(desc, ref):
    if ref:
        return ref
    for pat in (r"\bO\d{9,}\b", r"Cheque No[:\s]*\d+", r"CHQ\s*\d+", r"\b\d{9,}\b"):
        m = re.search(pat, desc, re.I)
        if m:
            return m.group(0)
    return ""

def _channel(desc):
    U = desc.upper()
    if "RTGS" in U:                              return "RTGS"
    if "CEFT" in U or "SLIPS" in U:              return "CEFT/SLIPS"
    if "CHEQUE" in U or "CHQ" in U:              return "Cheque"
    if "CASH" in U or "CRDLESDEP" in U:          return "Cash"
    if "LP EFT" in U:                            return "EFT"
    if "MOBILEBANKING" in U or "TRANSFER" in U:  return "Transfer"
    if "BROK" in U:                              return "Broker"
    return "Other"

def payer_clusters(unmatched):
    df = Counter(); parsed = []
    for u in unmatched:
        t = _payer_tokens(u["desc"]); parsed.append((u, t))
        for x in set(t):
            df[x] += 1
    def key_of(t):
        if not t:
            return None
        return tuple(sorted(sorted(set(t), key=lambda x: (df[x], -len(x)))[:2]))
    raw = defaultdict(lambda: {"n": 0, "v": 0.0, "refs": [], "names": [], "dates": [], "ch": Counter()})
    for u, t in parsed:
        k = key_of(t)
        if not k:
            continue
        c = raw[k]; c["n"] += 1; c["v"] += u["amount"] or 0
        br = _bankref(u["desc"], u.get("ref", ""))
        if br: c["refs"].append(br)
        c["names"].append(" ".join(t[:4])); c["ch"][_channel(u["desc"])] += 1
        if isinstance(u["date"], datetime):
            c["dates"].append(u["date"])
    keys = list(raw.keys())
    reps = {k: Counter(raw[k]["names"]).most_common(1)[0][0] for k in keys}
    parent = {k: k for k in keys}
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb: parent[rb] = ra
    block = defaultdict(list)
    for k in keys:
        for tk in k: block[tk].append(k)
    for tk, ks in block.items():
        for i in range(len(ks)):
            for j in range(i + 1, len(ks)):
                if fuzz.token_set_ratio(reps[ks[i]], reps[ks[j]]) >= 80:
                    union(ks[i], ks[j])
    M = defaultdict(lambda: {"n": 0, "v": 0.0, "refs": [], "names": [], "dates": [], "ch": Counter()})
    for k in keys:
        root = find(k); m, s = M[root], raw[k]
        m["n"] += s["n"]; m["v"] += s["v"]; m["refs"] += s["refs"]
        m["names"] += s["names"]; m["dates"] += s["dates"]; m["ch"] += s["ch"]
    clusters = []
    for k, c in M.items():
        rep = " ".join(w.capitalize() for w in Counter(c["names"]).most_common(1)[0][0].split())
        RU = rep.upper()
        if not rep.strip() or not any(len(t) >= 4 for t in rep.split()):
            continue
        if any(j in RU for j in ("DEPOSIT", "TRANSFER", "TRANSACTION", "CREDIT",
                                 "INSUAR", "INSURANC", "INSU ")):
            continue
        if RU.strip() in ("INSU", "INSUR", "CHEQUE", "CASH"):
            continue
        ds = sorted(c["dates"])
        clusters.append({"name": rep, "count": c["n"], "total": round(c["v"], 2),
                         "refs": list(dict.fromkeys(c["refs"])),
                         "first": ds[0] if ds else None, "last": ds[-1] if ds else None,
                         "channel": c["ch"].most_common(1)[0][0] if c["ch"] else ""})
    clusters.sort(key=lambda x: -x["total"])
    return clusters


# ==========================================================================
# Excel writing
# ==========================================================================
FONT = "Arial"; MONEY = "#,##0.00"
THIN = Side(style="thin", color="D9D9D9"); BD = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
def _hdr(ws, cols, widths, color="1F4E78"):
    ws.append(cols)
    for ci, w in enumerate(widths, 1):
        c = ws.cell(1, ci); c.fill = PatternFill("solid", fgColor=color)
        c.font = Font(name=FONT, bold=True, color="FFFFFF", size=10); c.border = BD
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(ci)].width = w

def write_workbook(ag, label, out, matched, additional, clusters, has_ref):
    c = ag.c
    wb = openpyxl.Workbook(); wb.remove(wb.active)

    # ---- Matched ----
    ws = wb.create_sheet("Matched")
    cols = ["Bank Date", "Bank Description", "Amount Paid", "Matched On", "Paid vs Invoice",
            "Policy No", "Division", "Customer / Account", "Insured Name", "Producer",
            "Doc No", "NIC", "Vehicle", "Original Amt", "Recd Before", "Outstanding Bal",
            "Ageing Days", "Score"]
    widths = [12, 40, 13, 20, 13, 20, 16, 24, 26, 20, 18, 14, 14, 13, 13, 14, 9, 8]
    _hdr(ws, cols, widths)
    for u, i, score, matched_on, am in sorted(matched, key=lambda x: -(x[0]["amount"] or 0)):
        r = ag.rows[i]; paid = u["amount"]; org = fnum(r[c["org"]])
        ws.append([u["date"], u["desc"], paid, matched_on, paid_vs_invoice(paid, org),
                   r[c["policy"]], r[c["division"]], r[c["customer"]], r[c["insured"]],
                   r[c["producer"]], r[c["doc"]], r[c["nic"]], r[c["vehicle"]],
                   org, fnum(r[c["recd"]]), fnum(r[c["bal"]]), r[c["ageing"]], score])
    _style_rows(ws, money_cols=(2, 13, 14, 15), date_col=0)

    # ---- Additional Matches ----
    ws2 = wb.create_sheet("Additional Matches")
    cols2 = ["Bank Date", "Bank Description", "Amount Paid", "Match Basis", "Matched Name Token",
             "Amount Agrees On", "Policy No", "Division", "Insured Name", "Customer / Account",
             "Producer", "Original Amt", "Outstanding Bal", "Ageing Days"]
    _hdr(ws2, cols2, [12, 40, 13, 14, 20, 20, 20, 16, 26, 24, 20, 13, 14, 9], color="7030A0")
    for u, i, basis, token, agrees in additional:
        r = ag.rows[i]
        ws2.append([u["date"], u["desc"], u["amount"], basis, token, agrees,
                    r[c["policy"]], r[c["division"]], r[c["insured"]], r[c["customer"]],
                    r[c["producer"]], fnum(r[c["org"]]), fnum(r[c["bal"]]), r[c["ageing"]]])
    _style_rows(ws2, money_cols=(2, 11, 12), date_col=0)

    # ---- Payer Clusters ----
    ws3 = wb.create_sheet("Payer Clusters")
    _hdr(ws3, ["Payer / Entity (from narrative)", "# Transfers", "Total Amount", "First Date",
               "Last Date", "Main Channel", "Bank References (for reconciliation)"],
         [40, 12, 18, 13, 13, 14, 60], color="2E7D32")
    for cl in clusters:
        ws3.append([cl["name"], cl["count"], cl["total"], cl["first"], cl["last"],
                    cl["channel"], "  |  ".join(cl["refs"])])
    for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
        for cell in row:
            cell.font = Font(name=FONT, size=9); cell.border = BD
            cell.alignment = Alignment(vertical="center")
        row[2].number_format = MONEY
        row[3].number_format = "yyyy-mm-dd"; row[4].number_format = "yyyy-mm-dd"
        row[6].alignment = Alignment(vertical="top", wrap_text=True)
        if (row[2].value or 0) >= 1_000_000:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="FFF9C4")
    ws3.freeze_panes = "A2"
    if ws3.max_row > 1:
        ws3.auto_filter.ref = f"A1:G{ws3.max_row}"

    # wb.save(f"{OUTPUT_DIR}/{out}")
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output

def _style_rows(ws, money_cols, date_col):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.font = Font(name=FONT, size=9); cell.border = BD
            cell.alignment = Alignment(vertical="center")
        if date_col is not None:
            row[date_col].number_format = "yyyy-mm-dd"
        for ci in money_cols:
            row[ci].number_format = MONEY
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"



def load_sheet(bank_file, sheet, layout):
    wb = openpyxl.load_workbook(bank_file, read_only=True, data_only=True)
    ws = wb[sheet]

    out = []

    for r in ws.iter_rows(min_row=layout["start"], values_only=True):
        amt = fnum(r[layout["amount"]]) if layout["amount"] < len(r) else None
        if amt is None:
            continue

        desc = str(r[layout["desc"]] or "") if layout["desc"] < len(r) else ""
        if not desc:
            continue

        ref = (
            str(r[layout["ref"]] or "")
            if 0 <= layout["ref"] < len(r)
            else ""
        )

        out.append({
            "date": r[layout["date"]],
            "desc": desc,
            "amount": amt,
            "ref": ref,
        })

    wb.close()
    return out


# ==========================================================================
# Main
# ==========================================================================
# def main():
#     print("Loading aging ...")
    
#     # st.write("AGING_FILE:", AGING_FILE)
#     # st.write("Type:", type(AGING_FILE))
#     ag = Aging(AGING_FILE, AGING_SHEET)
#     # st.write(f"  {len(ag.rows):,} aging rows indexed\n")

#     for sheet, layout in BANK_SHEET_LAYOUTS.items():
#         rows = load_sheet(BANK_FILE,sheet, layout)
#         matched, unmatched = [], []
#         for u in rows:
#             hit = confident_match(ag, u["desc"], u["amount"])
#             if hit:
#                 i, score, matched_on, am = hit
#                 matched.append((u, i, score, matched_on, am))
#             else:
#                 unmatched.append(u)
#         add = additional_matches(ag, unmatched)
#         already = {(str(u["date"]), u["desc"], u["amount"]) for u, *_ in add}
#         add += nic_plate_singles(ag, unmatched, already)
#         promoted = {(str(u["date"]), u["desc"], u["amount"]) for u, *_ in add}
#         still = [u for u in unmatched
#                  if (str(u["date"]), u["desc"], u["amount"]) not in promoted]
#         clusters = payer_clusters(still)
#         write_workbook(ag, layout["label"], layout["out"], matched, add, clusters,
#                        has_ref=(layout["ref"] >= 0))
#         print(f"{layout['label']:16}: {len(rows):>5} txns -> {len(matched):>4} matched, "
#               f"{len(add):>4} additional, {len(clusters):>4} payer clusters -> {layout['out']}")

# if __name__ == "__main__":
#     main()

def main():

    if "reports" not in st.session_state:
        st.session_state.reports = {}

    if BANK_FILE is None:
        st.warning("Please upload the Bank Statement.")
        return

    if AGING_FILE is None:
        st.warning("Please upload the Aging file.")
        return

    if st.button("Process"):

        with st.spinner("Generating reports..."):

            ag = Aging(AGING_FILE, AGING_SHEET)

            st.session_state.reports = {}

            for sheet, layout in BANK_SHEET_LAYOUTS.items():

                rows = load_sheet(BANK_FILE, sheet, layout)

                matched, unmatched = [], []

                for u in rows:
                    hit = confident_match(ag, u["desc"], u["amount"])

                    if hit:
                        i, score, matched_on, am = hit
                        matched.append((u, i, score, matched_on, am))
                    else:
                        unmatched.append(u)

                add = additional_matches(ag, unmatched)

                already = {
                    (str(u["date"]), u["desc"], u["amount"])
                    for u, *_ in add
                }

                add += nic_plate_singles(ag, unmatched, already)

                promoted = {
                    (str(u["date"]), u["desc"], u["amount"])
                    for u, *_ in add
                }

                still = [
                    u for u in unmatched
                    if (str(u["date"]), u["desc"], u["amount"]) not in promoted
                ]

                clusters = payer_clusters(still)

                excel_file = write_workbook(
                    ag,
                    layout["label"],
                    layout["out"],
                    matched,
                    add,
                    clusters,
                    has_ref=(layout["ref"] >= 0),
                )

                st.session_state.reports[layout["out"]] = excel_file

        st.success("Processing completed!")


    # Download section (inside main, outside Process button)
    if st.session_state.reports:

        st.subheader("Download Reports")

        for filename, excel in st.session_state.reports.items():

            st.download_button(
                label=f"Download {filename}",
                data=excel,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{filename}"
            )


if __name__ == "__main__":
    main()