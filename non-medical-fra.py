import streamlit as st
import oracledb
import pandas as pd
from datetime import datetime
from datetime import datetime, timedelta
from openpyxl import load_workbook

# import pdfplumber
import calendar
import os

from io import BytesIO

st.set_page_config(initial_sidebar_state="collapsed",
                   page_title= "FRA Report",
                   page_icon= "https://encrypted-tbn0.gstatic.com/images?q=tbn:ANd9GcT21JCb1UK7-E6bbE-HytHGkwH3g0i4LqItAg&s")


os.environ["LD_LIBRARY_PATH"] = "/home/orient/OracleClient/instantclient_23_26"
oracledb.init_oracle_client(
    lib_dir="/home/orient/OracleClient/instantclient_23_26"
)


months = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December"
]

st.title("FRA Monthly report")

# User chooses month number
choice = st.selectbox(
    "Choose a month number:",
    options=range(1, 13)
)

# Convert month number to month name
selected_month = months[choice - 1]

st.success(f"You selected month {choice}: {selected_month}")

start_date_chosen = datetime(2026, choice, 1)

# Last day of the selected month
last_day = calendar.monthrange(2026, choice)[1]
end_date_chosen = datetime(2026, choice, last_day)

current_year = 2026

# Start date: first day of selected month in previous year
start_date_sarry = datetime(current_year - 1, choice, 1)

# First day of selected month in current year
current_month_start = datetime(current_year, choice, 1)

# End date: last day of previous month in current year
end_date_sarry = current_month_start - timedelta(days=1)

current_year = datetime.today().year

first_date = datetime(current_year, 1, 1)

if choice == 12:
    end_date_selected = datetime(current_year + 1, 1, 1) - timedelta(days=1)
else:
    end_date_selected = datetime(current_year, choice + 1, 1) - timedelta(days=1)

individual_brokers = {
    "003000000008",
    "003000000078",
    "003000000012",
    "003000000018",
    "003000000037",
    "003000000093",
    "003000000100",
    "003000000109",
    "003000000111",
    "003000000113",
    "004000000115",
    "003000000141"
}

def get_broker_type(source_code,source_name):
    source_code = str(source_code)
    source_name = str(source_name or "")

    if "direct" in source_name.lower() :
        return "Direct Source"

    elif source_code in individual_brokers:
        return "Individuals Broker"

    elif source_code.startswith("00400000"):
        return "Individuals Broker"

    elif source_code.startswith("0030000"):
        return "Brokerage Company"

    else:
        return "Not defined"


wb = load_workbook("/home/orient/non-medical/temp_fra.xlsx")

ws = wb["non-life"]

# Engineering

conn = oracledb.connect(
    user="MIS",
    password="MIS",
    dsn="10.245.2.11:1521/orcl"
)

query = """
SELECT *
FROM NON_MEDICAL_REGISTER
WHERE POLH_APPR_DT >= TO_DATE('2025-01-01', 'YYYY-MM-DD')
  AND POLH_DEPT_CODE = '20'

"""

data = pd.read_sql(query, conn)



df = data.copy()


# ## Cleaning data




df["POLH_APPR_DT"] = pd.to_datetime(df["POLH_APPR_DT"]).dt.normalize()




# ## First Table 


f_month_policyonly = df [
    (df["POLH_APPR_DT"]>= start_date_chosen) & 
    (df["POLH_APPR_DT"]<= end_date_chosen) &
    (df["POLH_END_NO_IDX"]=='003') & 
    (df["POLH_END_NO"].fillna("").astype(str).str.strip() == "")
]



B18 = f_month_policyonly["POLH_NO"].count()


f_month_policyonly.loc[f_month_policyonly["POLH_STS"].eq("Cancelled"), "SUMSI"] *= -1





F18 = f_month_policyonly["SUMSI"].sum()


f_month_policyonly_all = df [
    (df["POLH_APPR_DT"]>= start_date_chosen) & 
    (df["POLH_APPR_DT"]<= end_date_chosen) &
    (df["POLH_END_NO_IDX"]=='003') 
]





G54 = f_month_policyonly_all['NETPREMIUMLC'].sum()



# In[673]:


D54 = f_month_policyonly_all["POLH_NO"].count()


f_month_policyonly_all = f_month_policyonly_all[f_month_policyonly_all["POLH_STS"] == 'Cancelled']

C54 = f_month_policyonly_all["POLH_NO"].count()



F54 = f_month_policyonly_all["NETPREMIUMLC"].sum()



B54 =D54+C54

E54=G54-F54

#  =================================================================================================================



# f_sarry_df = df [
# (df["POLH_APPR_DT"] >=start_date_sarry)
# & (df["POLH_APPR_DT"] <=end_date_sarry)
# ]

f_sarry_df = df[(df["POLH_TO_DT"] >= start_date_chosen)]



f_sarry_df["POLH_APPR_DT"].min()




C18 =  f_sarry_df["POLH_NO"].count()


G18 = f_sarry_df["SUMSI"].sum()



# ## Table 2



df= df[df["POLH_END_NO_IDX"]=='003']




f_period_df  = df[
(df["POLH_APPR_DT"] >= first_date) & 
(df["POLH_APPR_DT"] <= end_date_selected)&
(df["POLH_END_NO"].fillna("").astype(str).str.strip() == "")
]




D36 = f_period_df["POLH_NO"].count()




E36 = f_period_df["NETPREMIUMLC"].sum()





df["POLH_STS"].value_counts()




f_period_df_2  = df[
(df["POLH_APPR_DT"] >= first_date) & 
(df["POLH_APPR_DT"] <= end_date_selected)&
(df["POLH_STS"] == "Cancelled")
]




C36 = f_period_df_2["POLH_NO"].count()



F36 = f_period_df_2["NETPREMIUMLC"].sum()




B36 =D36+C36



G36 =E36+F36



# ## Table 3

# Done in Table 1

# ## Table 4




# f_month_policyonly["Broker Type"] = f_month_policyonly["SOURCE_CODE"].apply(get_broker_type)

f_month_policyonly["Broker Type"] = f_month_policyonly.apply(
    lambda row: get_broker_type(row["SOURCE_CODE"], row["SURCE_NAME"]),
    axis=1
)




# Direct Source
direct_df = f_month_policyonly[f_month_policyonly["Broker Type"] == "Direct Source"]

B108 = direct_df["NETPREMIUMLC"].sum()
F108 = direct_df["SUMSI"].sum()


# Individuals Broker
individual_df = f_month_policyonly[f_month_policyonly["Broker Type"] == "Individuals Broker"]

C108 = individual_df["NETPREMIUMLC"].sum()
G108 = individual_df["SUMSI"].sum()


# Brokerage Company
brokerage_df = f_month_policyonly[f_month_policyonly["Broker Type"] == "Brokerage Company"]

D108 = brokerage_df["NETPREMIUMLC"].sum()
H108 = brokerage_df["SUMSI"].sum()





# ## Filling the Template 

# In[711]:


# First table

ws["B18"] = B18
ws["C18"] = C18
ws["F18"] = F18
ws["G18"] = G18
#======================================================#

# Second table

ws["B36"] = B36
ws["C36"] = C36
ws["D36"] = D36
ws["E36"] = E36
ws["F36"] = F36
ws["G36"] = G36

#======================================================#

# Third table

ws["B54"] = B54
ws["C54"] = C54

ws["D54"] = D54
ws["E54"] = E54

ws["F54"] = F54
ws["G54"] = G54

#======================================================#

# Forth table

ws["B108"] = B108
ws["C108"] = C108
ws["D108"] = D108
ws["H108"] = H108
ws["F108"] = F108
ws["G108"] = G108
#======================================================#



# ## Save sheet

# In[715]:


# wb.save("Jun_EFRA_report_26-07-26_eng.xlsx")





# # Marine

# In[270]:


conn = oracledb.connect(
    user="MIS",
    password="MIS",
    dsn="10.245.2.11:1521/orcl"
)

query = """
SELECT *
FROM NON_MEDICAL_REGISTER
WHERE POLH_APPR_DT >= TO_DATE('2025-01-01', 'YYYY-MM-DD')
  AND POLH_DEPT_CODE IN ('50','51','52')
"""

data = pd.read_sql(query, conn)


# In[271]:


data.head(5)


# ## Cleaning data 

# In[505]:


df = data.copy()


# In[507]:


df["POLH_APPR_DT"] = pd.to_datetime(df["POLH_APPR_DT"]).dt.normalize()
df["POLH_TO_DT"] = pd.to_datetime(df["POLH_TO_DT"]).dt.normalize()


# ## Table 1 

# In[510]:


f_month_df  = df[
(df["POLH_APPR_DT"] >= start_date_chosen) & 
(df["POLH_APPR_DT"] <= end_date_chosen)
]


# In[512]:


f_month_df_pure =f_month_df[ f_month_df["POLH_STS"]!="Cancelled"]


# In[514]:


f_month_df.loc[f_month_df["POLH_STS"].eq("Cancelled"), "SUMSI"] *= -1


# In[516]:


lob_month_count = (
    f_month_df_pure
    .groupby("POLH_DEPT_CODE")
    .agg(   
        count=("POLH_NO", "size")
    )
)



lob_month_amount = (
    f_month_df
    .groupby("POLH_DEPT_CODE")
    .agg(   
        amount=("SUMSI", "sum")
    )
)


# In[518]:


B12 = lob_month_count.loc['50', "count"] if '50' in lob_month_count.index else 0
B13 = lob_month_count.loc['51', "count"] if '51' in lob_month_count.index else 0
B14 = lob_month_count.loc['52', "count"] if '52' in lob_month_count.index else 0

F12 = lob_month_amount.loc['50', "amount"] if '50' in lob_month_amount.index else 0
F13 = lob_month_amount.loc['51', "amount"] if '51' in lob_month_amount.index else 0
F14 = lob_month_amount.loc['52', "amount"] if '52' in lob_month_amount.index else 0



f_sarry_df = df [
# (df["POLH_APPR_DT"] >=start_date_sarry)&
# (df["POLH_APPR_DT"] <=end_date_sarry) & 
(~df["POLH_PROD_CODE"].isin(["50010", "50050"])) &
(df["POLH_TO_DT"] >= start_date_chosen)
]




f_sarry_df.loc[f_sarry_df["POLH_STS"].eq("Cancelled"), "SUMSI"] *= -1



lob_sarry_month = (
    f_sarry_df
    .groupby("POLH_DEPT_CODE") # Dept
    .agg(
        amount=("SUMSI", "sum"),
        count=("POLH_NO", "size")
    )
)




C12 = lob_sarry_month.loc['50', "count"] if '50' in lob_sarry_month.index else 0
C13 = lob_sarry_month.loc['51', "count"] if '51' in lob_sarry_month.index else 0
C14 = lob_sarry_month.loc['52', "count"] if '52' in lob_sarry_month.index else 0

G12 = lob_sarry_month.loc['50', "amount"] if '50' in lob_sarry_month.index else 0
G13 = lob_sarry_month.loc['51', "amount"] if '51' in lob_sarry_month.index else 0
G14 = lob_sarry_month.loc['52', "amount"] if '52' in lob_sarry_month.index else 0



# ## Table 2



f_period_df  = df[
(df["POLH_APPR_DT"] >= first_date) & 
(df["POLH_APPR_DT"] <= end_date_selected)
]


f_period_df_pure =f_period_df[ f_period_df["POLH_STS"]!="Cancelled"]



lob_period = (
    f_period_df_pure
    .groupby("POLH_DEPT_CODE")
    .agg(
        
        count=("POLH_NO", "size")
    )
)



lob_period_amount = (
    f_period_df
    .groupby("POLH_DEPT_CODE")
    .agg(
        
        amount=("NETPREMIUMLC", "sum")
    )
)



D30 = lob_period.loc['50', "count"] if '50' in lob_period.index else 0
D31 = lob_period.loc['51', "count"] if '51' in lob_period.index else 0
D32 = lob_period.loc['52', "count"] if '52' in lob_period.index else 0

E30 = lob_period_amount.loc['50', "amount"] if '50' in lob_period_amount.index else 0
E31 = lob_period_amount.loc['51', "amount"] if '51' in lob_period_amount.index else 0
E32 = lob_period_amount.loc['52', "amount"] if '52' in lob_period_amount.index else 0



f_period_df_cancel =f_period_df[ f_period_df["POLH_STS"]=="Cancelled"]



lob_period = (
    f_period_df_cancel
    .groupby("POLH_CLASS_CODE")
    .agg(
        amount=("NETPREMIUMLC", "sum"),
        count=("POLH_NO", "size")
    )
)



C30 = lob_period.loc['50', "count"] if '50' in lob_period.index else 0
C31 = lob_period.loc['51', "count"] if '51' in lob_period.index else 0
C32 = lob_period.loc['52', "count"] if '52' in lob_period.index else 0

F30 = lob_period.loc['50', "amount"] if '50' in lob_period.index else 0
F31 = lob_period.loc['51', "amount"] if '51' in lob_period.index else 0
F32 = lob_period.loc['52', "amount"] if '52' in lob_period.index else 0



B30 = D30+C30
B31 = D31+C31
B32 = D32+C32

G30 = E30-abs(F30)
G31 = E31-abs(F31)
G32 = E32-abs(F32)


# ## Table 3


f_month_df  = df[
(df["POLH_APPR_DT"] >= start_date_chosen) & 
(df["POLH_APPR_DT"] <= end_date_chosen)
]




f_month_df_pure =f_month_df[ f_month_df["POLH_STS"]!="Cancelled"]



lob_month = (
    f_month_df_pure
    .groupby("POLH_DEPT_CODE")
    .agg(
        
        count=("POLH_NO", "size")
    )
)


lob_month_amount = (
    f_month_df
    .groupby("POLH_DEPT_CODE")
    .agg(  
        amount=("NETPREMIUMLC", "sum")
    )
)




D48 = lob_month.loc['50', "count"] if '50' in lob_month.index else 0
D49 = lob_month.loc['51', "count"] if '51' in lob_month.index else 0
D50 = lob_month.loc['52', "count"] if '52' in lob_month.index else 0

E48 = lob_month_amount.loc['50', "amount"] if '50' in lob_month_amount.index else 0
E49 = lob_month_amount.loc['51', "amount"] if '51' in lob_month_amount.index else 0
E50 = lob_month_amount.loc['52', "amount"] if '52' in lob_month_amount.index else 0






f_month_df_cancel =f_month_df[ f_month_df["POLH_STS"]=="Cancelled"]



lob_month = (
    f_month_df_cancel
    .groupby("POLH_CLASS_CODE")
    .agg(
        amount=("NETPREMIUMLC", "sum"),
        count=("POLH_NO", "size")
    )
)




C48 = lob_month.loc['50', "count"] if '50' in lob_month.index else 0
C49 = lob_month.loc['51', "count"] if '51' in lob_month.index else 0
C50 = lob_month.loc['52', "count"] if '52' in lob_month.index else 0

F48 = lob_month.loc['50', "amount"] if '50' in lob_month.index else 0
F49 = lob_month.loc['51', "amount"] if '51' in lob_month.index else 0
F50 = lob_month.loc['52', "amount"] if '52' in lob_month.index else 0




B48 = D48+C48
B49 = D49+C49
B50 = D50+C50

G48 = E48-abs(F48)
G49 = E49-abs(F49)
G50 = E50-abs(F50)


# ## Table 4



# f_month_df["Broker Type"] = f_month_df["SOURCE_CODE"].apply(get_broker_type)



f_month_df["Broker Type"] = f_month_df.apply(
    lambda row: get_broker_type(row["SOURCE_CODE"], row["SURCE_NAME"]),
    axis=1
)



# Direct Source
direct_df = f_month_df[f_month_df["Broker Type"] == "Direct Source"]

lob_direct = (
    direct_df
    .groupby("POLH_CLASS_CODE")
    .agg(
        amount_n=("NETPREMIUMLC", "sum"),
        amount_s=("SUMSI", "sum")
    )
)

F102 = lob_direct.loc['50', "amount_s"] if '50' in lob_direct.index else 0
F103 = lob_direct.loc['51', "amount_s"] if '51' in lob_direct.index else 0
F104 = lob_direct.loc['52', "amount_s"] if '52' in lob_direct.index else 0

B102 = lob_direct.loc['50', "amount_n"] if '50' in lob_direct.index else 0
B103 = lob_direct.loc['51', "amount_n"] if '51' in lob_direct.index else 0
B104 = lob_direct.loc['52', "amount_n"] if '52' in lob_direct.index else 0



# Individuals Broker
individual_df = f_month_df[f_month_df["Broker Type"] == "Individuals Broker"]

lob_individual = (
    individual_df
    .groupby("POLH_CLASS_CODE")
    .agg(
        amount_n=("NETPREMIUMLC", "sum"),
        amount_s=("SUMSI", "sum")
    )
)

C102 = lob_individual.loc['50', "amount_n"] if '50' in lob_individual.index else 0
C103 = lob_individual.loc['51', "amount_n"] if '51' in lob_individual.index else 0
C104 = lob_individual.loc['52', "amount_n"] if '52' in lob_individual.index else 0

G102 = lob_individual.loc['50', "amount_s"] if '50' in lob_individual.index else 0
G103 = lob_individual.loc['51', "amount_s"] if '51' in lob_individual.index else 0
G104 = lob_individual.loc['52', "amount_s"] if '52' in lob_individual.index else 0



# Brokerage Company
brokerage_df = f_month_df[f_month_df["Broker Type"] == "Brokerage Company"]

lob_comp = (
    brokerage_df
    .groupby("POLH_CLASS_CODE")
    .agg(
        amount_n=("NETPREMIUMLC", "sum"),
        amount_s=("SUMSI", "sum")
    )
)

D102 = lob_comp.loc['50', "amount_n"] if '50' in lob_comp.index else 0
D103 = lob_comp.loc['51', "amount_n"] if '51' in lob_comp.index else 0
D104 = lob_comp.loc['52', "amount_n"] if '52' in lob_comp.index else 0

H102 = lob_comp.loc['50', "amount_s"] if '50' in lob_comp.index else 0
H103 = lob_comp.loc['51', "amount_s"] if '51' in lob_comp.index else 0
H104 = lob_comp.loc['52', "amount_s"] if '52' in lob_comp.index else 0


# ## Filling in sheet


# First table

ws["B12"] = B12
ws["B13"] = B13
ws["B14"] = B14

ws["C12"] = C12
ws["C13"] = C13
ws["C14"] = C14

ws["F12"] = F12
ws["F13"] = F13
ws["F14"] = F14

ws["G12"] = G12
ws["G13"] = G13
ws["G14"] = G14
#======================================================#

# Second table

ws["B30"] = B30
ws["B31"] = B31
ws["B32"] = B32

ws["C30"] = C30
ws["C31"] = C31
ws["C32"] = C32

ws["D30"] = D30
ws["D31"] = D31
ws["D32"] = D32

ws["E30"] = E30
ws["E31"] = E31
ws["E32"] = E32

ws["F30"] = F30
ws["F31"] = F31
ws["F32"] = F32

ws["G30"] = G30
ws["G31"] = G31
ws["G32"] = G32

#======================================================#

# Third table

ws["B48"] = B48
ws["B49"] = B49
ws["B50"] = B50

ws["C48"] = C48
ws["C49"] = C49
ws["C50"] = C50

ws["D48"] = D48
ws["D49"] = D49
ws["D50"] = D50

ws["E48"] = E48
ws["E49"] = E49
ws["E50"] = E50

ws["F48"] = F48
ws["F49"] = F49
ws["F50"] = F50

ws["G48"] = G48
ws["G49"] = G49
ws["G50"] = G50

#======================================================#

# Forth table

ws["B102"] = B102
ws["B103"] = B103
ws["B104"] = B104

ws["C102"] = C102
ws["C103"] = C103
ws["C104"] = C104

ws["D102"] = D102
ws["D103"] = D103
ws["D104"] = D104


ws["F102"] = F102
ws["F103"] = F103
ws["F104"] = F104

ws["G102"] = G102
ws["G103"] = G103
ws["G104"] = G104

ws["H102"] = H102
ws["H103"] = H103
ws["H104"] = H104
#======================================================#


# ## Save file



# wb.save("JUN_EFRA_report_27-07-26_marine_3.xlsx")


# # General Accidents


conn = oracledb.connect(
    user="MIS",
    password="MIS",
    dsn="10.245.2.11:1521/orcl"
)

query = """
SELECT *
FROM NON_MEDICAL_REGISTER
WHERE POLH_APPR_DT >= TO_DATE('2025-01-01', 'YYYY-MM-DD')
  AND POLH_DEPT_CODE = '30'
"""

data = pd.read_sql(query, conn)


# ## Cleaning data 



df= data.copy()




df["POLH_APPR_DT"] = pd.to_datetime(df["POLH_APPR_DT"]).dt.normalize()
df["POLH_TO_DT"] = pd.to_datetime(df["POLH_TO_DT"]).dt.normalize()


# ## Table #1




f_month_df  = df[
(df["POLH_APPR_DT"] >= start_date_chosen) & 
(df["POLH_APPR_DT"] <= end_date_chosen)
]


f_month_df_new =f_month_df[ f_month_df["POLH_STS"]=="New Policy"]
f_month_df_cancel =f_month_df[ f_month_df["POLH_STS"]=="Cancelled"]


count_new = f_month_df_new["POLH_NO"].count()
count_cancel = f_month_df_cancel["POLH_NO"].count()

B20 = count_new-count_cancel


f_month_df_new_endo = f_month_df[ f_month_df["POLH_STS"].isin(["New Policy","Endorsed"])]


new_endo_si = f_month_df_new_endo["SUMSI"].sum()
cancel_si = f_month_df_cancel["SUMSI"].sum()

F20 = new_endo_si-cancel_si


f_sarry_df = df[(df["POLH_TO_DT"] >= start_date_chosen)]


f_sarry_df_count = f_sarry_df[f_sarry_df["POLH_STS"].isin(["New Policy","Renewed"])]
C20 = f_sarry_df_count["POLH_NO"].count()




f_sarry_df_si = f_sarry_df[f_sarry_df["POLH_STS"]!="Cancelled"]
G20 = f_sarry_df_si["SUMSI"].sum()



# ## Table #2


f_period_df  = df[
(df["POLH_APPR_DT"] >= first_date) & 
(df["POLH_APPR_DT"] <= end_date_selected)
]

f_period_new_count = f_period_df[f_period_df["POLH_STS"].isin(["New Policy","Renewed"])]
B38 = f_period_new_count["POLH_NO"].count()



f_period_cancel_count = f_period_df[f_period_df["POLH_STS"]=="Cancelled"]
C38 = f_period_cancel_count["POLH_NO"].count()



D38=B38-C38


F38 = f_period_cancel_count["NETPREMIUMLC"].sum()


# ## Table #3



B56 = f_month_df.loc[f_month_df["POLH_STS"] == "New Policy", "POLH_NO"].count()


C56 = f_month_df.loc[f_month_df["POLH_STS"] == "Cancelled", "POLH_NO"].count()


D56=B56-C56



# In[455]:


E56 = f_month_df.loc[f_month_df["POLH_STS"].isin( ["New Policy","Endorsed"]), "NETPREMIUMLC"].sum()



# In[457]:


F56 = f_month_df.loc[f_month_df["POLH_STS"] == "Cancelled", "NETPREMIUMLC"].sum()



# In[459]:


G56=E56+F56



# ## Table #4

# In[650]:


# f_month_df["Broker Type"] = f_month_df["SOURCE_CODE"].apply(get_broker_type)
f_month_df.loc[f_month_df["POLH_STS"].eq("Cancelled"), "SUMSI"] *= -1

f_month_df["Broker Type"] = f_month_df.apply(
    lambda row: get_broker_type(row["SOURCE_CODE"], row["SURCE_NAME"]),
    axis=1
)


# In[652]:


# Direct Source
direct_df = f_month_df[f_month_df["Broker Type"] == "Direct Source"]

B110 = direct_df["NETPREMIUMLC"].sum()
F110 = direct_df["SUMSI"].sum()


# Individuals Broker
individual_df = f_month_df[f_month_df["Broker Type"] == "Individuals Broker"]

C110 = individual_df["NETPREMIUMLC"].sum()
G110 = individual_df["SUMSI"].sum()


# Brokerage Company
brokerage_df = f_month_df[f_month_df["Broker Type"] == "Brokerage Company"]

D110 = brokerage_df["NETPREMIUMLC"].sum()
H110 = brokerage_df["SUMSI"].sum()




# First table

ws["B20"] = B20
ws["C20"] = C20
ws["F20"] = F20
ws["G20"] = G20
#======================================================#

# Second table

ws["B38"] = B38
ws["C38"] = C38
ws["D38"] = D38
# ws["E38"] = E38
ws["F38"] = F38
# ws["G38"] = G38

#======================================================#

# Third table

ws["B56"] = B56
ws["C56"] = C56

ws["D56"] = D56
ws["E56"] = E56

ws["F56"] = F56
ws["G56"] = G56

#======================================================#

# Forth table

ws["B110"] = B110
ws["C110"] = C110
ws["D110"] = D110
ws["H110"] = H110
ws["F110"] = F110
ws["G110"] = G110
#======================================================#



# In[524]:


# wb.save("Jun_EFRA_report_GA_22-07-26.xlsx")


# # Property 



conn = oracledb.connect(
    user="MIS",
    password="MIS",
    dsn="10.245.2.11:1521/orcl"
)

query = """
SELECT *
FROM NON_MEDICAL_REGISTER
WHERE POLH_APPR_DT >= TO_DATE('2025-01-01', 'YYYY-MM-DD')
  AND POLH_DEPT_CODE = '10'
"""

data = pd.read_sql(query, conn)


df=data.copy()



df["POLH_APPR_DT"] = pd.to_datetime(df["POLH_APPR_DT"]).dt.normalize()
df["POLH_TO_DT"] = pd.to_datetime(df["POLH_TO_DT"]).dt.normalize()




# ## Table #1



f_month_df  = df[
(df["POLH_APPR_DT"] >= start_date_chosen) & 
(df["POLH_APPR_DT"] <= end_date_chosen)
]


f_month_df_new =f_month_df[f_month_df["POLH_STS"].isin(["New Policy","Renewed"])]
# f_month_df_cancel =f_month_df[ f_month_df["POLH_STS"]=="Cancelled"]





count_new = f_month_df_new["POLH_NO"].count()
# count_cancel = f_month_df_cancel["POLH_NO"].count()

# B11 = count_new-count_cancel
B11=count_new



f_month_df_new_endo = f_month_df[ f_month_df["POLH_STS"]!="Cancelled"]
f_month_df_cancel = f_month_df[ f_month_df["POLH_STS"]=="Cancelled"]




new_endo_si = f_month_df_new_endo["SUMSI"].sum()
cancel_si = f_month_df_cancel["SUMSI"].sum()

F11 = new_endo_si-cancel_si



f_sarry_df = df[(df["POLH_TO_DT"] >= start_date_chosen)]

f_sarry_df_count = f_sarry_df[f_sarry_df["POLH_STS"].isin(["New Policy","Renewed"])]
C11 = f_sarry_df_count["POLH_NO"].count()


f_sarry_df_si = f_sarry_df[f_sarry_df["POLH_STS"]!="Cancelled"]
f_sarry_df_si_cancel = f_sarry_df[f_sarry_df["POLH_STS"]=="Cancelled"]

si_pure = f_sarry_df_si["SUMSI"].sum()
si_cancel = f_sarry_df_si_cancel["SUMSI"].sum()
G11= si_pure-si_cancel



# ## Table #2

# In[265]:


f_period_df  = df[
(df["POLH_APPR_DT"] >= first_date) & 
(df["POLH_APPR_DT"] <= end_date_selected)
]


f_period_new_count = f_period_df[f_period_df["POLH_STS"].isin(["New Policy","Renewed"])]
B29 = f_period_new_count["POLH_NO"].count()



f_period_cancel_count = f_period_df[f_period_df["POLH_STS"]=="Cancelled"]
C29 = f_period_cancel_count["POLH_NO"].count()



# In[215]:


D29=B29-C29



F29 = f_period_cancel_count["NETPREMIUMLC"].sum()



E29 = f_period_df["NETPREMIUMLC"].sum()



# ## Table #3



B47 = f_month_df.loc[f_month_df["POLH_STS"] == "New Policy", "POLH_NO"].count()



# In[226]:


C47 = f_month_df.loc[f_month_df["POLH_STS"] == "Cancelled", "POLH_NO"].count()



D47=B47-C47



E47 = f_month_df.loc[f_month_df["POLH_STS"].isin( ["New Policy","Endorsed"]), "NETPREMIUMLC"].sum()



F47 = f_month_df.loc[f_month_df["POLH_STS"] == "Cancelled", "NETPREMIUMLC"].sum()


G47=E47+F47



# ## Table #4

# In[238]:


f_month_df  = df[
(df["POLH_APPR_DT"] >= start_date_chosen) & 
(df["POLH_APPR_DT"] <= end_date_chosen)
]





# f_month_df["Broker Type"] = f_month_df["SOURCE_CODE"].apply(get_broker_type)

f_month_df["Broker Type"] = f_month_df.apply(
    lambda row: get_broker_type(row["SOURCE_CODE"], row["SURCE_NAME"]),
    axis=1
)




# In[248]:


# Direct Source
direct_df = f_month_df[f_month_df["Broker Type"] == "Direct Source"]

B101 = direct_df["NETPREMIUMLC"].sum()
F101 = direct_df["SUMSI"].sum()


# Individuals Broker
individual_df = f_month_df[f_month_df["Broker Type"] == "Individuals Broker"]

C101 = individual_df["NETPREMIUMLC"].sum()
G101 = individual_df["SUMSI"].sum()


# Brokerage Company
brokerage_df = f_month_df[f_month_df["Broker Type"] == "Brokerage Company"]

D101 = brokerage_df["NETPREMIUMLC"].sum()
H101 = brokerage_df["SUMSI"].sum()




# ## Filling template

# In[256]:


# First table

ws["B11"] = B11
ws["C11"] = C11
ws["F11"] = F11
ws["G11"] = G11
#======================================================#

# Second table

ws["B29"] = B29
ws["C29"] = C29
ws["D29"] = D29
# ws["E29"] = E29
ws["F29"] = F29
# ws["G29"] = G29

#======================================================#

# Third table

ws["B47"] = B47
ws["C47"] = C47

ws["D47"] = D47
ws["E47"] = E47

ws["F47"] = F47
ws["G47"] = G47

#======================================================#

# Forth table

ws["B101"] = B101
ws["C101"] = C101
ws["D101"] = D101
ws["H101"] = H101
ws["F101"] = F101
ws["G101"] = G101
#======================================================#



# ## save report

# In[259]:


# wb.save("Jun_EFRA_report_property_27-07-26.xlsx")


output = BytesIO()

# Save the modified workbook into memory
wb.save(output)

# Important: move to the beginning of the file
output.seek(0)

st.download_button(
    label="📥 Download Resulted Report",
    data=output,
    file_name="Jun_EFRA_report_27-07-26.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# #################################################################################################################

